"""
ImBox Lead Tracker
===================
Läser ringlistan, kollar Gmail efter skickade prospekteringsmail
och uppdaterar Excel med status.

Genererar en statusrapport + lista med leads som behöver uppföljning.

Användning via bash_tool:
  python3 lead_tracker.py <excel_fil> <gmail_search_json>

gmail_search_json skapas av Claude via gmail_search_messages och sparas hit.
"""
import json, sys, os
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

FOLLOWUP_DAYS = 5  # Dagar innan uppföljning föreslås


def read_sent_status(gmail_results_path):
    """Läs Gmail-sökresultat (JSON sparad av Claude)."""
    with open(gmail_results_path, encoding="utf-8") as f:
        return json.load(f)


def read_excel_leads(excel_path):
    """Läs leads från Excel."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Ringlista"]
    headers = [cell.value for cell in ws[1]]
    leads = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        lead = dict(zip(headers, row))
        lead["_row"] = row_idx
        leads.append(lead)
    wb.close()
    return leads


def match_leads_to_emails(leads, gmail_threads):
    """Matcha leads med Gmail-trådar baserat på e-postadress."""
    email_status = {}
    for thread in gmail_threads:
        to_addr = thread.get("to", "").lower().strip()
        has_reply = thread.get("has_reply", False)
        sent_date = thread.get("sent_date", "")
        thread_id = thread.get("thread_id", "")

        if to_addr:
            email_status[to_addr] = {
                "has_reply": has_reply,
                "sent_date": sent_date,
                "thread_id": thread_id,
                "days_since": 0,
            }
            if sent_date:
                try:
                    sent = datetime.fromisoformat(sent_date[:10])
                    email_status[to_addr]["days_since"] = (datetime.now() - sent).days
                except:
                    pass

    matched = []
    for lead in leads:
        email = (lead.get("E-post") or "").lower().strip()
        if email in email_status:
            lead["_mail_status"] = email_status[email]
        else:
            lead["_mail_status"] = None
        matched.append(lead)

    return matched


def generate_report(matched_leads):
    """Generera statusrapport."""
    report = {
        "total": len(matched_leads),
        "mail_sent": 0,
        "replied": 0,
        "needs_followup": [],
        "no_reply_recent": [],
        "not_contacted": [],
        "meetings_booked": 0,
    }

    for lead in matched_leads:
        status = lead.get("Status", "Ej kontaktad")
        ms = lead.get("_mail_status")

        if status == "Möte/demo bokat":
            report["meetings_booked"] += 1
        elif ms:
            report["mail_sent"] += 1
            if ms["has_reply"]:
                report["replied"] += 1
            elif ms["days_since"] >= FOLLOWUP_DAYS:
                report["needs_followup"].append({
                    "company": lead.get("Bolagsnamn", ""),
                    "email": lead.get("E-post", ""),
                    "days_since": ms["days_since"],
                    "chat_provider": lead.get("Chattleverantör", ""),
                    "thread_id": ms.get("thread_id", ""),
                    "row": lead.get("_row", 0),
                })
            else:
                report["no_reply_recent"].append({
                    "company": lead.get("Bolagsnamn", ""),
                    "days_since": ms["days_since"],
                })
        else:
            if status == "Ej kontaktad":
                report["not_contacted"].append({
                    "company": lead.get("Bolagsnamn", ""),
                    "email": lead.get("E-post", ""),
                    "chat_provider": lead.get("Chattleverantör", ""),
                })

    return report


def generate_followup_body(lead, attempt=2):
    """Generera uppföljningsmailtext."""
    company = lead.get("company", "ert bolag")
    kontakt = lead.get("contact_name", "")
    chat = lead.get("chat_provider", "")

    greeting = f"Hej {kontakt}," if kontakt else "Hej,"

    if attempt == 2:
        body = f"""{greeting}

Jag hör av mig igen angående kundkommunikation hos {company}. Jag förstår att det kan vara svårt att hinna med allt — ville bara dubbelkolla om mitt förra mail hamnade rätt.

Kort sammanfattat erbjuder vi på ImBox en svensk plattform för livechatt, chatbot, ticketing och telefoni. Snabb setup, svensk support.

Skulle 15 minuter nästa vecka fungera för en snabb demo?

Vänliga hälsningar,
ImBox"""
    else:
        body = f"""{greeting}

Sista försöket — jag vill inte störa, men ville ge er möjligheten innan jag stänger ner kontakten.

Om kundkommunikation är en fråga för {company} längre fram, finns vi på imbox.io. Vi hjälper gärna till med en kostnadsfri analys av er nuvarande setup.

Allt gott,
ImBox"""

    return body


def print_report(report):
    """Skriv ut rapport till terminal."""
    print("\n" + "=" * 50)
    print("📊 LEAD-TRACKER RAPPORT")
    print("=" * 50)
    print(f"  Totalt leads:        {report['total']}")
    print(f"  Mail skickade:       {report['mail_sent']}")
    print(f"  Fått svar:           {report['replied']}")
    print(f"  Möten bokade:        {report['meetings_booked']}")
    print(f"  Ej kontaktade:       {len(report['not_contacted'])}")
    print()

    if report["needs_followup"]:
        print(f"⚠️  BEHÖVER UPPFÖLJNING ({len(report['needs_followup'])} st):")
        for f in report["needs_followup"]:
            print(f"    • {f['company']} — {f['days_since']} dagar sedan")
            if f["chat_provider"]:
                print(f"      Har: {f['chat_provider']}")
        print()

    if report["no_reply_recent"]:
        print(f"⏳  Inväntar svar ({len(report['no_reply_recent'])} st):")
        for r in report["no_reply_recent"]:
            print(f"    • {r['company']} — {r['days_since']} dag(ar)")
    print()

    # Spara rapport som JSON
    output_path = "/home/claude/tracker_report.json"
    with open(output_path, "w", encoding="utf-8") as fp:
        json.dump(report, fp, ensure_ascii=False, indent=2, default=str)
    print(f"📁 Rapport sparad: {output_path}")

    # Spara uppföljningar för Claude att agera på
    if report["needs_followup"]:
        followups = []
        for f in report["needs_followup"]:
            followups.append({
                "to": f["email"],
                "company": f["company"],
                "chat_provider": f["chat_provider"],
                "thread_id": f["thread_id"],
                "days_since": f["days_since"],
                "body": generate_followup_body(f, attempt=2),
                "subject": None,  # Sätts via threadId (reply)
            })
        followup_path = "/home/claude/pending_followups.json"
        with open(followup_path, "w", encoding="utf-8") as fp:
            json.dump(followups, fp, ensure_ascii=False, indent=2)
        print(f"📧 {len(followups)} uppföljningar förberedda: {followup_path}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Användning: python3 lead_tracker.py <excel_fil> <gmail_search_json>")
        print()
        print("gmail_search_json = JSON-fil med Gmail-sökresultat")
        print("Skapas av Claude genom att köra gmail_search_messages")
        sys.exit(1)

    excel_path = sys.argv[1]
    gmail_path = sys.argv[2]

    leads = read_excel_leads(excel_path)
    gmail_data = read_sent_status(gmail_path)
    matched = match_leads_to_emails(leads, gmail_data)
    report = generate_report(matched)
    print_report(report)
