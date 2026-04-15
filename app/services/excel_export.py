from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

STATUS_VALS = [
    "Ej kontaktad", "Ringt – inget svar", "Mail skickat",
    "Återkoppling bokad", "Möte/demo bokat", "Tackat nej", "Fel kontaktperson",
]
STATUS_FARGER = {
    "Ej kontaktad": "F2F2F2",
    "Ringt – inget svar": "FFE699",
    "Mail skickat": "BDD7EE",
    "Återkoppling bokad": "C6EFCE",
    "Möte/demo bokat": "70AD47",
    "Tackat nej": "FF7575",
    "Fel kontaktperson": "D9D9D9",
}
MX_FARGER = {
    "ok": ("E2EFDA", "548235"),
    "catch_all": ("FFF2CC", "9C5700"),
    "no_mx": ("FF7575", "9C0006"),
    "invalid": ("D9D9D9", "5F5E5A"),
    "error": ("D9D9D9", "5F5E5A"),
}
MX_DISPLAY = {"ok": "✓ OK", "catch_all": "~ info@", "no_mx": "✗ Ingen MX", "invalid": "✗ Ogiltig", "error": "? DNS-fel"}


def build_workbook(campaign, leads):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ringlista"

    headers = ["#", "Bolagsnamn", "Org.nr", "Omsättning", "Anställda", "Bransch", "SNI",
               "Ort", "Hemsida", "E-post", "MX", "Kontaktperson", "Kontaktroll",
               "LinkedIn-koll", "Telefon", "Status", "Datum", "Kommentar"]
    col_widths = [4, 35, 14, 14, 10, 28, 8, 16, 30, 30, 12, 25, 25, 16, 16, 22, 13, 35]

    hfill = PatternFill("solid", start_color="1F3864")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    afill = PatternFill("solid", start_color="EBF0FA")
    bdr = Border(bottom=Side(style="thin", color="D0D7E8"))

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    dv = DataValidation(
        type="list", formula1='"' + ",".join(STATUS_VALS) + '"',
        allow_blank=True, showErrorMessage=True,
    )
    ws.add_data_validation(dv)

    mx_prio = {"ok": 0, "catch_all": 1, "error": 2, "no_mx": 3, "invalid": 4}
    sorted_leads = sorted(leads, key=lambda l: (mx_prio.get(l.mx_status, 3), l.company_name or ""))

    for ri, l in enumerate(sorted_leads, 2):
        fill = afill if ri % 2 == 0 else None
        mx_display = MX_DISPLAY.get(l.mx_status, "?")
        vals = [ri - 1, l.company_name, l.org_nr, l.revenue, l.employees, l.industry, l.sni,
                l.city, l.website, l.email, mx_display, l.contact_name, l.contact_role,
                "Kolla →" if l.linkedin_url else "", l.phone,
                l.call_status or "Ej kontaktad",
                l.call_date.isoformat() if l.call_date else "",
                l.comment or ""]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=col, value=val)
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(vertical="center")
            c.border = bdr
            if fill:
                c.fill = fill

        if l.website:
            cell = ws.cell(row=ri, column=9)
            cell.hyperlink = l.website
            cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

        mx_cell = ws.cell(row=ri, column=11)
        bg, fg = MX_FARGER.get(l.mx_status, ("D9D9D9", "5F5E5A"))
        mx_cell.fill = PatternFill("solid", start_color=bg)
        mx_cell.font = Font(name="Arial", size=10, color=fg, bold=(l.mx_status == "no_mx"))
        mx_cell.alignment = Alignment(horizontal="center", vertical="center")

        li_cell = ws.cell(row=ri, column=14)
        if l.linkedin_url:
            li_cell.hyperlink = l.linkedin_url
            li_cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

        status_cell = ws.cell(row=ri, column=16)
        status_fill = STATUS_FARGER.get(l.call_status or "Ej kontaktad", "F2F2F2")
        status_cell.fill = PatternFill("solid", start_color=status_fill)
        status_cell.font = Font(name="Arial", size=10)
        status_cell.alignment = Alignment(horizontal="center", vertical="center")
        dv.add(status_cell)

    ws.freeze_panes = "B2"
    if sorted_leads:
        ws.auto_filter.ref = f"A1:R{len(sorted_leads) + 1}"

    info = wb.create_sheet("Info")
    info["A1"] = "Pipefire — Kampanj"
    info["A1"].font = Font(bold=True, size=14, name="Arial")
    for row, (label, val) in enumerate([
        ("Namn:", campaign.name),
        ("SNI-prefix:", campaign.sni_prefix),
        ("Min omsättning:", f"{campaign.min_msek} MSEK"),
        ("Storleksklass:", f"{campaign.min_klass}–{campaign.max_klass}"),
        ("Antal bolag:", len(sorted_leads)),
        ("Exporterad:", date.today().isoformat()),
    ], 3):
        info.cell(row=row, column=1, value=label).font = Font(bold=True, name="Arial")
        info.cell(row=row, column=2, value=val).font = Font(name="Arial")
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 40

    return wb
