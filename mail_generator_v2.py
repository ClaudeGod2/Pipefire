"""
ImBox Mail-generator v2
========================
Läser en berikad ringlista (Excel) och genererar personaliserade
mailutkast. Filtrerar bort leads med dålig MX eller bytt jobb.

Användning:
  python3 mail_generator.py /path/to/ringlista.xlsx [antal_leads]
"""
import json, sys, os
import openpyxl

AVSÄNDARNAMN = "Adam"
MAX_LEADS = 10


def read_leads(excel_path, max_leads=MAX_LEADS):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["Ringlista"]
    headers = [cell.value for cell in ws[1]]
    leads = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(leads) >= max_leads:
            break
        lead = dict(zip(headers, row))
        # Filtrering
        if not lead.get("E-post"):
            continue
        mx = lead.get("MX", "")
        if "Ingen MX" in str(mx) or "Ogiltig" in str(mx):
            continue
        if lead.get("Status") and lead["Status"] != "Ej kontaktad":
            continue
        leads.append(lead)
    wb.close()
    return leads


def generate_subject(lead):
    company = lead.get("Bolagsnamn", "ert bolag")
    return f"Kundkommunikation hos {company} — en tanke"


def generate_body(lead):
    company = lead.get("Bolagsnamn", "ert bolag")
    kontakt = lead.get("Kontaktperson", "")
    bransch = lead.get("Bransch", "")
    ort = lead.get("Ort", "")

    greeting = f"Hej {kontakt}," if kontakt else "Hej,"

    # Bygg en personlig touch baserat på tillgänglig data
    context_line = ""
    if bransch:
        context_line = f"Allt fler företag inom {bransch.lower()} investerar i digital kundkommunikation"
    else:
        context_line = "Allt fler svenska företag investerar i digital kundkommunikation"

    body = f"""{greeting}

Jag kontaktar er på {company} för att {context_line} — och jag tänkte att det kanske är relevant även för er.

ImBox är en svensk plattform som samlar livechatt, chatbot, ticketing och telefoni i ett enda gränssnitt. Det innebär att era kunder kan nå er på det sätt de föredrar, och att ert team slipper hoppa mellan verktyg.

Tre saker som brukar avgöra:
- Snabb setup — de flesta är live på under en vecka
- Svensk support och hosting (GDPR)
- Mätbar effekt — våra kunder ser i snitt 30% kortare svarstider

Är det värt 15 minuter av er tid att se hur det fungerar?

Vänliga hälsningar,
{AVSÄNDARNAMN}
ImBox"""

    return body


def generate_drafts(leads):
    drafts = []
    for lead in leads:
        drafts.append({
            "to": lead.get("E-post", ""),
            "subject": generate_subject(lead),
            "body": generate_body(lead),
            "company": lead.get("Bolagsnamn", ""),
            "contact_name": lead.get("Kontaktperson", ""),
            "mx_status": lead.get("MX", ""),
            "row_number": lead.get("#", ""),
        })
    return drafts


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Användning: python3 mail_generator.py <excel_fil> [antal_leads]")
        sys.exit(1)

    excel_path = sys.argv[1]
    max_leads = int(sys.argv[2]) if len(sys.argv) > 2 else MAX_LEADS

    if not os.path.exists(excel_path):
        print(f"Fil hittades inte: {excel_path}")
        sys.exit(1)

    leads = read_leads(excel_path, max_leads)
    print(f"Läste {len(leads)} leads (MX OK, ej kontaktade)")

    drafts = generate_drafts(leads)

    output_path = "/home/claude/pending_drafts.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(drafts, f, ensure_ascii=False, indent=2)

    print(f"\n📧 {len(drafts)} mailutkast förberedda → {output_path}")

    if drafts:
        d = drafts[0]
        print(f"\n── Preview: {d['company']} ──")
        print(f"   Till: {d['to']}")
        print(f"   Ämne: {d['subject']}")
        print(f"   MX: {d['mx_status']}")
        print(f"   Längd: {len(d['body'])} tecken")
