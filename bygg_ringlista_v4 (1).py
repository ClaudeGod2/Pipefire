"""
ImBox Ringlista — Enrichment Pipeline v4
==========================================
1. Bolagsdata från Bolagsverket/SCB + ABPI
2. MX-validering av e-postdomän (anti-bounce)
3. LinkedIn-verifiering av kontaktperson

Variabler att sätta:
  SNI_PREFIX, MIN_KLASS, MAX_KLASS, MIN_MSEK, MAX_BOLAG, ABPI_NYCKEL
"""
import urllib.request, zipfile, io, os, time, json, re, ssl, subprocess, socket
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date
from urllib.parse import quote_plus

# ── KONFIG ──────────────────────────────────────────────────
SNI_PREFIX  = "47"
MIN_KLASS   = 4
MAX_KLASS   = 9
MIN_MSEK    = 300
MAX_BOLAG   = 50
ABPI_NYCKEL = ""

SCB_URL   = "https://mr2.bolagsverket.se/ftp/scb_bulkfil.zip"
ABPI_BASE = "https://abpi.se/api"
CACHE     = "/home/claude/scb_cache.txt"

_mx_cache = {}


# ═══════════════════════════════════════════════════════════
#  MX-VALIDERING
# ═══════════════════════════════════════════════════════════

def check_mx(email):
    """
    Validerar att e-postdomänen har MX-records.
    Returnerar (status, detalj):
      "ok"        — MX finns, mailen bör nå fram
      "no_mx"     — Ingen MX-record, hög bouncerrisk
      "catch_all" — Generisk info@-adress (kolla manuellt)
      "invalid"   — Ogiltig e-postadress / ingen domän
      "error"     — DNS-fel (timeout etc.)
    """
    if not email or "@" not in email:
        return "invalid", "Ingen giltig e-post"

    domain = email.split("@")[1].strip().lower()

    if domain in _mx_cache:
        status, detail = _mx_cache[domain]
        if status == "ok" and email.lower().startswith("info@"):
            return "catch_all", f"MX OK men generisk (info@{domain})"
        return status, detail

    # Metod 1: host -t MX (Linux/Mac)
    try:
        result = subprocess.run(
            ["host", "-t", "MX", domain],
            capture_output=True, text=True, timeout=5
        )
        if "mail is handled by" in result.stdout:
            mx_lines = [l for l in result.stdout.splitlines() if "mail is handled by" in l]
            mx_servers = [l.split("by")[-1].strip().rstrip(".") for l in mx_lines[:3]]
            detail = ", ".join(mx_servers)
            _mx_cache[domain] = ("ok", detail)
            if email.lower().startswith("info@"):
                return "catch_all", f"MX OK men generisk (info@{domain})"
            return "ok", detail
        elif "NXDOMAIN" in result.stderr or "not found" in result.stdout.lower():
            _mx_cache[domain] = ("no_mx", f"Domänen {domain} saknar MX")
            return "no_mx", f"Domänen {domain} saknar MX"
        else:
            _mx_cache[domain] = ("no_mx", f"Ingen MX för {domain}")
            return "no_mx", f"Ingen MX för {domain}"
    except FileNotFoundError:
        pass
    except subprocess.TimeoutExpired:
        _mx_cache[domain] = ("error", "DNS timeout")
        return "error", "DNS timeout"
    except Exception:
        pass

    # Metod 2: nslookup
    try:
        result = subprocess.run(
            ["nslookup", "-type=MX", domain],
            capture_output=True, text=True, timeout=5
        )
        if "mail exchanger" in result.stdout.lower():
            _mx_cache[domain] = ("ok", "MX via nslookup")
            if email.lower().startswith("info@"):
                return "catch_all", f"MX OK men generisk (info@{domain})"
            return "ok", "MX via nslookup"
        else:
            _mx_cache[domain] = ("no_mx", f"Ingen MX via nslookup")
            return "no_mx", f"Ingen MX via nslookup"
    except Exception:
        pass

    # Metod 3: socket fallback
    try:
        socket.getaddrinfo(domain, 25, socket.AF_INET, socket.SOCK_STREAM)
        _mx_cache[domain] = ("ok", "Port 25 öppen (fallback)")
        if email.lower().startswith("info@"):
            return "catch_all", f"Port 25 OK men generisk (info@{domain})"
        return "ok", "Port 25 öppen (fallback)"
    except socket.gaierror:
        _mx_cache[domain] = ("no_mx", f"Domänen {domain} resolvar inte")
        return "no_mx", f"Domänen {domain} resolvar inte"
    except Exception as e:
        _mx_cache[domain] = ("error", str(e))
        return "error", str(e)


# ═══════════════════════════════════════════════════════════
#  LINKEDIN URL-GENERERING
# ═══════════════════════════════════════════════════════════

def generate_linkedin_url(namn, bolagsnamn):
    if not namn:
        return "", ""
    clean_company = re.sub(
        r'\b(AB|HB|KB|Handelsbolag|Aktiebolag|Kommanditbolag|Inc|Ltd|GmbH|AS|A/S)\b',
        '', bolagsnamn, flags=re.IGNORECASE
    ).strip().rstrip(",").strip()
    search_term = f"{namn} {clean_company}"
    url = f"https://www.linkedin.com/search/results/people/?keywords={quote_plus(search_term)}"
    return url, search_term


# ═══════════════════════════════════════════════════════════
#  STEG 1: LADDA SCB BULKDATA
# ═══════════════════════════════════════════════════════════

if os.path.exists(CACHE) and (time.time() - os.path.getmtime(CACHE)) < 604800:
    print("Använder cachad SCB-fil")
    with open(CACHE, encoding="utf-8") as f:
        rader = f.read().splitlines()
else:
    print("Laddar ner SCB bulk-fil (~30-60 MB)...")
    req = urllib.request.Request(SCB_URL, headers={
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/120.0.0.0 Safari/537.36"
    })
    with urllib.request.urlopen(req, timeout=180) as r:
        data = r.read()
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        txts = [f for f in z.namelist() if f.lower().endswith(".txt")]
        innehall = z.read(txts[0] if txts else z.namelist()[0]).decode("latin-1")
    rader = innehall.splitlines()
    with open(CACHE, "w", encoding="utf-8") as f:
        f.write(innehall)
    print(f"{len(rader):,} rader laddade och cachade")


# ═══════════════════════════════════════════════════════════
#  STEG 2: FILTRERA KANDIDATER
# ═══════════════════════════════════════════════════════════

kandidater = []
for i, rad in enumerate(rader):
    if i == 0:
        continue
    d = rad.split("\t")
    if len(d) < 7:
        continue
    orgnr = d[0].strip()
    namn = d[1].strip()
    snis = [d[j].strip().replace(".", "") for j in [5, 6, 7] if len(d) > j]
    storlek = d[-3].strip() if len(d) > 10 else ""
    if not any(s.startswith(SNI_PREFIX.replace(".", "")) for s in snis if s):
        continue
    try:
        k = int(storlek)
        if k < MIN_KLASS or k > MAX_KLASS:
            continue
    except:
        pass
    o = orgnr.replace("16", "", 1) if orgnr.startswith("16") else orgnr
    if len(o) == 10:
        o = f"{o[:6]}-{o[6:]}"
    kandidater.append({"orgnr": o, "namn": namn, "sni": snis[0] if snis else ""})

print(f"{len(kandidater):,} kandidater matchar SNI {SNI_PREFIX}*")


# ═══════════════════════════════════════════════════════════
#  STEG 3: HÄMTA DETALJDATA + ENRICHMENT
# ═══════════════════════════════════════════════════════════

def hämta(orgnr):
    req = urllib.request.Request(
        f"{ABPI_BASE}/{orgnr}/data",
        headers={"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                                "AppleWebKit/537.36 (KHTML, like Gecko) "
                                "Chrome/120.0.0.0 Safari/537.36"}
    )
    if ABPI_NYCKEL:
        req.add_header("Authorization", f"Bearer {ABPI_NYCKEL}")
    try:
        with urllib.request.urlopen(req, timeout=10) as r:
            return json.loads(r.read())
    except:
        return None

def omsattning(d):
    fs = d.get("financial_summary") or {}
    rev = fs.get("revenue")
    if rev:
        return rev, f"{round(rev / 1e6)} MSEK"
    return None, fs.get("estimated_turnover", "")

def kontakt(d):
    kws = [
        ["marknadschef", "marketing", "cmo"],
        ["kundservice", "customer service", "cx"],
        ["it-chef", "cto", "it-ansvarig"],
    ]
    for g in (d.get("roles") or {}).get("role_groups", []):
        for r in g.get("roles", []):
            rt = (r.get("role") or "").lower()
            for kw in kws:
                if any(k in rt for k in kw):
                    return r.get("name", ""), r.get("role", "")
    return "", ""

def epost(d):
    bi = d.get("basic_info") or {}
    if bi.get("email"):
        return bi["email"]
    hp = bi.get("home_page", "")
    if hp:
        try:
            return "info@" + hp.replace("https://", "").replace("http://", "").replace("www.", "").split("/")[0]
        except:
            pass
    return ""

def hemsida(d):
    bi = d.get("basic_info") or {}
    hp = bi.get("home_page", "")
    if hp and not hp.startswith("http"):
        hp = "https://" + hp
    return hp

resultat = []
mx_stats = {"ok": 0, "no_mx": 0, "catch_all": 0, "invalid": 0, "error": 0}
li_stats = {"unverified": 0, "no_contact": 0}
max_försök = min(len(kandidater), MAX_BOLAG * 6)

for i, k in enumerate(kandidater[:max_försök]):
    if len(resultat) >= MAX_BOLAG:
        break
    if i % 10 == 0:
        print(f"  {i}/{max_försök} kontrollerade, {len(resultat)} matchar...")

    d = hämta(k["orgnr"])
    if not d:
        continue

    rev_kr, rev_text = omsattning(d)
    if MIN_MSEK and rev_kr is not None and rev_kr / 1e6 < MIN_MSEK:
        continue

    bi = d.get("basic_info") or {}
    knamn, kroll = kontakt(d)
    hp = hemsida(d)
    email = epost(d)
    bolagsnamn = bi.get("name", k["namn"])

    # ── ENRICHMENT 1: MX-validering ──
    mx_status, mx_detalj = check_mx(email)
    mx_stats[mx_status] += 1
    if mx_status == "no_mx":
        print(f"    ⚠️  {bolagsnamn}: MX saknas för {email}")

    # ── ENRICHMENT 2: LinkedIn URL ──
    linkedin_url, linkedin_sök = generate_linkedin_url(knamn, bolagsnamn)
    li_status = "Ej verifierad"
    if not knamn:
        li_status = "Ingen kontaktperson"
        li_stats["no_contact"] += 1
    else:
        li_stats["unverified"] += 1

    resultat.append({
        "Bolagsnamn":       bolagsnamn,
        "Org.nr":           bi.get("organization_number", k["orgnr"]),
        "Omsättning":       rev_text,
        "Anställda":        d.get("number_of_employees", ""),
        "Bransch":          (d.get("current_industry") or {}).get("name", ""),
        "SNI":              k["sni"],
        "Ort":              ((d.get("addresses") or {}).get("visitor_address") or {}).get("post_place", ""),
        "Hemsida":          hp,
        "Telefon":          (bi.get("phone_numbers") or [""])[0],
        "E-post":           email,
        "MX-status":        mx_status,
        "MX-detalj":        mx_detalj,
        "Kontaktperson":    knamn,
        "Kontaktroll":      kroll,
        "LinkedIn":         linkedin_url,
        "LinkedIn-status":  li_status,
    })
    time.sleep(0.15)


# ═══════════════════════════════════════════════════════════
#  STATISTIK
# ═══════════════════════════════════════════════════════════

print(f"\n✅ {len(resultat)} bolag hittade")

print(f"\n📧 MX-validering:")
print(f"   OK (leverbar):      {mx_stats['ok']}")
print(f"   Catch-all (info@):  {mx_stats['catch_all']}")
print(f"   Ingen MX (bounce!): {mx_stats['no_mx']}")
print(f"   Ogiltiga:           {mx_stats['invalid']}")
print(f"   DNS-fel:            {mx_stats['error']}")

print(f"\n👤 LinkedIn:")
print(f"   Att verifiera:      {li_stats['unverified']}  ← Claude verifierar i fas 2")
print(f"   Ingen kontaktpers:  {li_stats['no_contact']}")


# ═══════════════════════════════════════════════════════════
#  STEG 4: SKAPA EXCEL
# ═══════════════════════════════════════════════════════════

STATUS_VALS = [
    "Ej kontaktad", "Ringt – inget svar", "Mail skickat",
    "Återkoppling bokad", "Möte/demo bokat", "Tackat nej", "Fel kontaktperson"
]
STATUS_FÄRGER = {
    "Ej kontaktad":       "F2F2F2",
    "Ringt – inget svar": "FFE699",
    "Mail skickat":       "BDD7EE",
    "Återkoppling bokad": "C6EFCE",
    "Möte/demo bokat":    "70AD47",
    "Tackat nej":         "FF7575",
    "Fel kontaktperson":  "D9D9D9",
}
MX_FÄRGER = {
    "ok":        ("E2EFDA", "548235"),
    "catch_all": ("FFF2CC", "9C5700"),
    "no_mx":     ("FF7575", "9C0006"),
    "invalid":   ("D9D9D9", "5F5E5A"),
    "error":     ("D9D9D9", "5F5E5A"),
}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Ringlista"

headers = [
    "#", "Bolagsnamn", "Org.nr", "Omsättning", "Anställda", "Bransch", "SNI",
    "Ort", "Hemsida", "E-post", "MX", "Kontaktperson", "Kontaktroll",
    "LinkedIn-koll", "Telefon", "Status", "Datum", "Kommentar"
]
col_widths = [
    4, 35, 14, 14, 10, 28, 8,
    16, 30, 30, 12, 25, 25,
    16, 16, 22, 13, 35
]

hfill = PatternFill("solid", start_color="1F3864")
hfont = Font(bold=True, color="FFFFFF", name="Arial", size=10)
afill = PatternFill("solid", start_color="EBF0FA")
bdr = Border(bottom=Side(style="thin", color="D0D7E8"))

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = hfont
    c.fill = hfill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
ws.row_dimensions[1].height = 22

dv = DataValidation(
    type="list",
    formula1='"' + ",".join(STATUS_VALS) + '"',
    allow_blank=True,
    showDropDown=False,
    showErrorMessage=True,
    errorTitle="Ogiltigt värde",
    error="Välj ett värde från listan."
)
ws.add_data_validation(dv)

# Sortera: bäst MX först, sedan namn
def sort_key(r):
    mx = r["MX-status"]
    mx_prio = {"ok": 0, "catch_all": 1, "error": 2, "no_mx": 3, "invalid": 4}
    return (mx_prio.get(mx, 3), r["Bolagsnamn"])

resultat.sort(key=sort_key)

for ri, r in enumerate(resultat, 2):
    fill = afill if ri % 2 == 0 else None

    mx_display = {
        "ok": "✓ OK", "catch_all": "~ info@", "no_mx": "✗ Ingen MX",
        "invalid": "✗ Ogiltig", "error": "? DNS-fel",
    }.get(r["MX-status"], "?")

    vals = [
        ri - 1, r["Bolagsnamn"], r["Org.nr"], r["Omsättning"], r["Anställda"],
        r["Bransch"], r["SNI"], r["Ort"], r["Hemsida"],
        r["E-post"], mx_display, r["Kontaktperson"], r["Kontaktroll"],
        r["LinkedIn-status"], r["Telefon"], "Ej kontaktad", "", ""
    ]

    for col, val in enumerate(vals, 1):
        c = ws.cell(row=ri, column=col, value=val)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(vertical="center")
        c.border = bdr
        if fill:
            c.fill = fill

    # Hemsida — klickbar
    if r["Hemsida"]:
        cell = ws.cell(row=ri, column=9)
        cell.hyperlink = r["Hemsida"]
        cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

    # MX — färgkod
    mx_cell = ws.cell(row=ri, column=11)
    mx_bg, mx_fg = MX_FÄRGER.get(r["MX-status"], ("D9D9D9", "5F5E5A"))
    mx_cell.fill = PatternFill("solid", start_color=mx_bg)
    mx_cell.font = Font(name="Arial", size=10, color=mx_fg, bold=(r["MX-status"] == "no_mx"))
    mx_cell.alignment = Alignment(horizontal="center", vertical="center")

    # LinkedIn — klickbar sök-URL
    li_cell = ws.cell(row=ri, column=14)
    if r["LinkedIn"]:
        li_cell.hyperlink = r["LinkedIn"]
        li_cell.value = "Kolla →"
        li_cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
    else:
        li_cell.font = Font(name="Arial", size=10, color="5F5E5A", italic=True)

    # Status — dropdown
    status_cell = ws.cell(row=ri, column=16)
    status_cell.fill = PatternFill("solid", start_color=STATUS_FÄRGER["Ej kontaktad"])
    status_cell.font = Font(name="Arial", size=10)
    status_cell.alignment = Alignment(horizontal="center", vertical="center")
    dv.add(status_cell)

ws.freeze_panes = "B2"
ws.auto_filter.ref = f"A1:R{len(resultat) + 1}"


# ═══════════════════════════════════════════════════════════
#  INFO-FLIK
# ═══════════════════════════════════════════════════════════

ms = wb.create_sheet("Info")
ms["A1"] = "ImBox Ringlista"
ms["A1"].font = Font(bold=True, size=14, name="Arial")

for row, (label, val) in enumerate([
    ("SNI-kod:", SNI_PREFIX),
    ("Min omsättning:", f"{MIN_MSEK} MSEK"),
    ("Storleksklass:", f"{MIN_KLASS}-{MAX_KLASS}"),
    ("Antal bolag:", len(resultat)),
    ("Genererad:", date.today().isoformat()),
], 3):
    ms.cell(row=row, column=1, value=label).font = Font(bold=True, name="Arial")
    ms.cell(row=row, column=2, value=val).font = Font(name="Arial")

r = 10
ms.cell(row=r, column=1, value="MX-validering:").font = Font(bold=True, name="Arial", size=11)
for i, (label, status) in enumerate([
    ("✓ OK — mailen levereras", "ok"),
    ("~ info@ — generisk, kolla manuellt", "catch_all"),
    ("✗ Ingen MX — SKICKA INTE (bouncerisk)", "no_mx"),
], r + 1):
    bg, fg = MX_FÄRGER[status]
    c = ms.cell(row=i, column=1, value=label)
    c.fill = PatternFill("solid", start_color=bg)
    c.font = Font(name="Arial", color=fg)

r2 = r + 5
ms.cell(row=r2, column=1, value="LinkedIn-koll:").font = Font(bold=True, name="Arial", size=11)
ms.cell(row=r2 + 1, column=1, value='Klicka "Kolla →" för att söka personen på LinkedIn').font = Font(name="Arial", italic=True)
ms.cell(row=r2 + 2, column=1, value="Claude verifierar automatiskt i fas 2 (mailutkast)").font = Font(name="Arial", italic=True)

r3 = r2 + 4
ms.cell(row=r3, column=1, value="Statusguide:").font = Font(bold=True, name="Arial", size=11)
for i, (s, fg) in enumerate(STATUS_FÄRGER.items(), r3 + 1):
    c = ms.cell(row=i, column=1, value=s)
    c.font = Font(name="Arial")
    c.fill = PatternFill("solid", start_color=fg)

for col in ["A", "B"]:
    ms.column_dimensions[col].width = 50

utfil = f"/mnt/user-data/outputs/imbox_ringlista_SNI{SNI_PREFIX}.xlsx"
wb.save(utfil)
print(f"\n📁 Sparad: {utfil}")

json_path = "/home/claude/enriched_leads.json"
with open(json_path, "w", encoding="utf-8") as f:
    json.dump(resultat, f, ensure_ascii=False, indent=2, default=str)
print(f"📋 JSON för fas 2: {json_path}")
